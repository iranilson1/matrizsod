from typing import Optional, Tuple, Union
import customtkinter as ctk
from tkinter import *
from tkinter.ttk import *
from tkinter import messagebox
import openpyxl, xlrd
import pathlib
from openpyxl import Workbook
from PIL import Image
import pandas as pd 

class Backend():   
    def salvando(self):
        arquivo = pathlib.Path('sistemaEscola.xlsx')
        if arquivo.exists():
            pass
        else:
            #FOLHA DA FUNÇÃO SISTEMA
            arquivo = Workbook()
            folha1 = arquivo.active
            folha1.title = 'Sistema'
            folha1['A1'] = 'CODIGO'
            folha1['B1'] = 'NOME'
            
            #FOLHA DA FUNÇÃO PERFIL DO SISTEMA 
            folha2=arquivo.create_sheet('perfilSistema')
            folha2['A1'] = 'CODIGO'
            folha2['B1'] = 'NOME'
            folha2['C1'] = 'DESCRIÇÃO'

            #FOLHA DA FUNÇÃO MATRIZSOD 
            folha3=arquivo.create_sheet('matrizSOD')
            folha3['A1'] = 'CODIGO1'
            folha3['B1'] = 'NOME1'
            folha3['C1'] = 'CODIGO2'
            folha3['D1'] = 'NOME2'

            #FOLHA DA FUNÇÃO perfiluser
            folha4=arquivo.create_sheet('PerfilUser')
            folha4['A1'] = 'CPF'
            folha4['B1'] = 'CODIGO'
            folha4['C1'] = 'NOME'
            arquivo.save(r'sistemaEscola.xlsx')
            
    def salvaSistema(self):
        #pegar os daos que estão no formulario do sistema 
        self.codigo = self.codigoSistemas.get()
        self.sistema = self.nomeSistemas.get()

        # Carregue o arquivo Excel em um DataFrame
        dataframesistema = pd.read_excel('.\sistemaEscola.xlsx', sheet_name='Sistema')

        # Verifique a unicidade dos valores na coluna
        codigo_duplicados = dataframesistema.loc[dataframesistema['CODIGO']==self.codigo,'CODIGO']
        nome_duplicados = dataframesistema.loc[dataframesistema['NOME']==self.sistema,'NOME']
        

        #salvar os dados na folha do excel
        if(self.codigo=='' or self.sistema==''):
            messagebox.showerror('sistema','ERRO\n Por favor prencha todos os campos')
        elif(not (list(codigo_duplicados) or list(nome_duplicados))):
            arquivo = openpyxl.load_workbook(r'sistemaEscola.xlsx')
            folha1 = arquivo.get_sheet_by_name(r'Sistema')
            folha1.cell(column=1, row=folha1.max_row+1, value=self.codigo)
            folha1.cell(column=2, row=folha1.max_row, value=self.sistema)
            arquivo.save(r'sistemaEscola.xlsx')
            msg = messagebox.showinfo(title='Estado do cadastro', message= "Parabens! serviço cadastrado com sucesso")

            #apagando o texto das entrys
            #self.codigoSistemas.set('')
            #self.nomeSistemas.set('')
        else:
            messagebox.showerror('sistema','ERRO\n Codigo ou nome ja existentes, verifique a lista cadastrada')

    def salvaPerfilServico(self):
        #pegar os daos que estão no formulario do sistema 
        self.codigo = self.codigo_perfil.get()
        self.sistema = self.nome_perfil.get()
        self.caixa = self.r_caixaTexto.get('0.0', 'end')

        # Carregue o arquivo Excel em um DataFrame
        dataframesistema = pd.read_excel('.\sistemaEscola.xlsx', sheet_name='perfilSistema')

        # Verifique a unicidade dos valores na coluna
        nome_duplicados = dataframesistema.loc[dataframesistema['NOME']==self.sistema,'NOME']

        #salvar os dados na folha do excel
        if(self.codigo=='' or self.sistema=='' or self.caixa==''):
            messagebox.showerror('sistema','ERRO\n Por favor prencha todos os campos')
        elif(not list(nome_duplicados)):
            arquivo = openpyxl.load_workbook(r'sistemaEscola.xlsx')
            folha2 = arquivo.get_sheet_by_name(r'perfilSistema')
            folha2.cell(column=1, row=folha2.max_row+1, value=self.codigo)
            folha2.cell(column=2, row=folha2.max_row, value=self.sistema)
            folha2.cell(column=3, row=folha2.max_row, value=self.caixa)
            arquivo.save(r'sistemaEscola.xlsx')
            msg = messagebox.showinfo(title='Estado do cadastro', message= "Parabens! perfil do serviço cadastrado com sucesso")

            #apagando o texto das entrys
            self.codigo_perfil.set('')
            self.nome_perfil.set('')
            self.r_caixaTexto.delete('0.0','end')
        else:
            messagebox.showerror('sistema','ERRO\n Nome ja existentes, verifique a lista cadastrada')
    
    def salvaMatriz(self):
        #pegar os dados que estão no formulario do sistema 
        self.codigo1 = self.codigo_sistema_1.get()
        self.sistema1 = self.nome_sistema_1.get()
        self.codigo2 = self.codigo_sistema_2.get()
        self.sistema2 = self.nome_sistema_2.get()

        #salvar os dados na folha do excel
        if(self.codigo1=='' or self.sistema1=='' or self.codigo2=='' or self.sistema2==''):
            messagebox.showerror('sistema','ERRO\n Por favor selecione todos os campos')
        else:
            arquivo = openpyxl.load_workbook(r'sistemaEscola.xlsx')
            folha3 = arquivo.get_sheet_by_name(r'matrizSOD')
            folha3.cell(column=1, row=folha3.max_row+1, value=self.codigo1)
            folha3.cell(column=2, row=folha3.max_row, value=self.sistema1)
            folha3.cell(column=3, row=folha3.max_row, value=self.codigo2)
            folha3.cell(column=4, row=folha3.max_row, value=self.sistema2)
            arquivo.save(r'sistemaEscola.xlsx')
            msg = messagebox.showinfo(title='Estado do cadastro', message= "Parabens! Matriz de conflito cadastrado com sucesso")

            #apagando o texto das entrys
    def salvarUser(self):
        #pegar os dados que estão no formulario do sistema 
        self.cpfs = int(self.cpf.get())
        self.codigo = self.codigo_sistema.get()
        self.nome = self.nome_sistema.get()

        #primeiro vou buscar no banco de dados se esse cpf ja esta cadastrado e qual perfil esta cadastrado 
        dataframeMatriz1 = pd.read_excel('.\sistemaEscola.xlsx', sheet_name='PerfilUser')
        nomematriz2 = dataframeMatriz1.loc[dataframeMatriz1['CPF']==self.cpfs,'NOME']

        # Verifique a unicidade dos valores na coluna
        cpf_duplicados = dataframeMatriz1.loc[dataframeMatriz1['CPF']==self.cpfs,'CPF']

        #apos isso busco todos os conflitos que o sistema digitado tem
        dataframeMatriz2 = pd.read_excel('.\sistemaEscola.xlsx', sheet_name='matrizSOD')
        nomematriz3 = dataframeMatriz2.loc[dataframeMatriz2['NOME1']==self.nome,'NOME2']
        nomematriz4 = dataframeMatriz2.loc[dataframeMatriz2['NOME2']==self.nome,'NOME1']
        conflito=list(nomematriz3)+list(nomematriz4)
        #apos ter uma lista de perfis conflitantes, eu verifico se essa lista tem algum argumento constando nos ja cadastrados
        erro=0
        if(self.cpfs=='' or self.codigo=='' or self.nome==''):
            messagebox.showerror('sistema','ERRO\n Por favor selecione todos os campos')    
        elif(not list(cpf_duplicados)):
            for i in conflito:
                if i in  list(nomematriz2) :
                    messagebox.showerror('sistema','ERRO\n Perfil conflitante com um ja cadastrado')
                    erro = 1
                    break
            if erro == 0:
                #salvar os dados na folha do excel
                arquivo = openpyxl.load_workbook(r'sistemaEscola.xlsx')
                folha4 = arquivo.get_sheet_by_name(r'PerfilUser')
                folha4.cell(column=1, row=folha4.max_row+1, value=self.cpfs)
                folha4.cell(column=2, row=folha4.max_row, value=self.codigo)
                folha4.cell(column=3, row=folha4.max_row, value=self.nome)
                arquivo.save(r'sistemaEscola.xlsx')
                msg = messagebox.showinfo(title='Estado do cadastro', message= "Parabens! Perfil de usuario cadastrado com sucesso")

                #apagando o texto das entrys
                #self.cpf.set('')
                self.codigo_sistema.set('')
                self.nome_sistema.set('')
        else:
            messagebox.showerror('sistema', 'ERRO\n CPF ja cadastrado, veja a lista de cpfs cadastrados' )        

class App(ctk.CTk, Backend):
    def __init__(self):
        super().__init__()
        self.tema()
        self.tela()
        self.salvando()
        self.tela_inicial()
    
    def tema(self):
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("dark-blue")

    def tela(self):
        self.title("Sistema Escolar")
        self.geometry("700x500")
        self.resizable(False,False)

    def tela_inicial(self):
        # dentro de inicial frame
        #img = ctk.CTkImage(Image.open(r'C:\Users\irani\SynologyDrive\fullstack\trabalhos\matrizSOD\PROJETOMATRIZSOD\icon.jpg'), size=(300,200))
        #img = ctk.CTkImage(Image.open(r'C:\Users\irani\SynologyDrive\fullstack\trabalhos\matrizSOD\PROJETOMATRIZSOD\Gestao-escolar.jpg'), size=(700,200))
        #label_img = ctk.CTkLabel(self, image=img, text='').place(x=0,y=10)
        titulo = ctk.CTkLabel(self, text = 'Sistema de gestão escolar', font=('Century Gothic bold',24), text_color='#fff').place(x=200,y=10)
        #subtitulo = ctk.CTkLabel(self, text='Por favor, preencha todos os campos do formulario', font=('Century Gothic bold',12), text_color='#fff').place(x=0,y=40)
        
        #frame
        inicial_frame = ctk.CTkFrame(master=self, width=700, height= 400)
        inicial_frame.pack(side=RIGHT)

        # sistema completo com todas as telas 
        def tela_sistemas():
            #remover tela inicial
            inicial_frame.pack_forget()

            #criando tela de cadastro de sistema
            sistema_frame = ctk.CTkFrame(master=self, width=700, height= 400)
            sistema_frame.pack(side=RIGHT)

            titulo = ctk.CTkLabel(master=sistema_frame, text = 'Cadastre os sistemas', font=('Century Gothic bold',16), text_color='gray').place(x=20,y=10)
            label_codigo = ctk.CTkLabel(master=sistema_frame, text = 'Digite o codigo do sistema', font=('Century Gothic bold',16), text_color='#fff').place(x=265,y=65)
            self.codigoSistemas = ctk.CTkEntry(master=sistema_frame,placeholder_text= 'CDG', width=300)
            self.codigoSistemas.place(x=200, y=100)
            
            label_sistema = ctk.CTkLabel(master=sistema_frame, text = 'Digite o nome do sistema', font=('Century Gothic bold',16), text_color='#fff').place(x=265,y=140)
            self.nomeSistemas = ctk.CTkEntry(master=sistema_frame,placeholder_text= 'Sistema', width=300)
            self.nomeSistemas.place(x=200, y=175)
            
            def back():
                #removendo frame
                sistema_frame.pack_forget()

                #devolvendo frame da tela inicial
                inicial_frame.pack(side=RIGHT)
                
            voltar = ctk.CTkButton(master=sistema_frame, text='VOLTAR', font=('Century Gothic bold',16), text_color='#fff',command= back ).place(x=20, y=350)
            self.salvar = ctk.CTkButton(master=sistema_frame, text='SALVAR', font=('Century Gothic bold',16), text_color='#fff', fg_color='green',hover_color="#014B05", command= self.salvaSistema).place(x=545, y=350)
        def tela_perfil():
            #remover tela inicial
            inicial_frame.pack_forget()

            #criando tela de cadastro de sistema
            perfil_frame = ctk.CTkFrame(master=self, width=700, height= 400)
            perfil_frame.pack(side=RIGHT)

            #pegar um datafreme
            dataframe = pd.read_excel('.\sistemaEscola.xlsx', sheet_name='Sistema')
            valor = dataframe.loc[:,'CODIGO']
            

            titulo = ctk.CTkLabel(master=perfil_frame, text = 'Cadastro do perfil de acesso', font=('Century Gothic bold',16), text_color='gray').place(x=20,y=10)
            label_codigo_perfil = ctk.CTkLabel(master=perfil_frame, text = 'Digite o código do sistema', font=('Century Gothic bold',16), text_color='#fff').place(x=265,y=65)
            self.codigo_perfil = ctk.CTkComboBox(master=perfil_frame, values=list(valor))
            self.codigo_perfil.place(x=280, y=100)
            
            label_nome = ctk.CTkLabel(master=perfil_frame, text = 'Nome do perfil', font=('Century Gothic bold',16), text_color='#fff').place(x=295,y=140)
            self.nome_perfil = ctk.CTkEntry(master=perfil_frame,placeholder_text= 'Nome', width=300)
            self.nome_perfil.place(x=200, y=175)

            label_nome = ctk.CTkLabel(master=perfil_frame, text = 'Descrição', font=('Century Gothic bold',16), text_color='#fff').place(x=300,y=215)
            self.r_caixaTexto = ctk.CTkTextbox(master=perfil_frame, width=300, height=120, corner_radius=0)
            self.r_caixaTexto.place(x=200, y=250)

            def back():
                #removendo frame
                perfil_frame.pack_forget()

                #devolvendo frame da tela inicial
                inicial_frame.pack(side=RIGHT)
            
            voltar = ctk.CTkButton(master=perfil_frame, text='VOLTAR', font=('Century Gothic bold',16), text_color='#fff',command= back ).place(x=20, y=350)
            self.salvar = ctk.CTkButton(master=perfil_frame, text='SALVAR', font=('Century Gothic bold',16), text_color='#fff', fg_color='green',hover_color="#014B05", command= self.salvaPerfilServico ).place(x=545, y=350)
            
        def tela_matriz():
            #remover tela inicial
            inicial_frame.pack_forget()

            #criando tela de cadastro de sistema
            matriz_frame = ctk.CTkFrame(master=self, width=700, height= 400)
            matriz_frame.pack(side=RIGHT)

            #recuperando os valores do banco de dados
            dataframe = pd.read_excel('.\sistemaEscola.xlsx', sheet_name='Sistema')
            COD = dataframe.loc[:,'CODIGO']
            
            titulo = ctk.CTkLabel(master=matriz_frame, text = 'Cadastre dos conflitos', font=('Century Gothic bold',16), text_color='gray').place(x=20,y=10)

            labem_Ma_codigo_1 = ctk.CTkLabel(master=matriz_frame, text = 'Escolha o primeiro codigo do sistema ', font=('Century Gothic bold',16), text_color='#fff').place(x=220,y=35)
            self.nome_sistema_1 = ctk.CTkComboBox(master=matriz_frame,values=[''])
            self.nome_sistema_1.place(x=270, y=147)
            def combobox_callback(choice):
                if (choice):
                    #FILTRANDO OS PERFIS DE CADA SISTEMA
                    dataframeMatriz1 = pd.read_excel('.\sistemaEscola.xlsx', sheet_name='perfilSistema')
                    nomematriz1 = dataframeMatriz1.loc[dataframeMatriz1['CODIGO']==choice,'NOME']
                    self.nome_sistema_1 = ctk.CTkComboBox(master=matriz_frame,values=list(nomematriz1))
                    self.nome_sistema_1.place(x=270, y=147)
            label_sistema_1 = ctk.CTkLabel(master=matriz_frame, text = 'Escolha o perfil do sistema 1 ', font=('Century Gothic bold',16), text_color='#fff').place(x=252,y=110)            
            self.codigo_sistema_1 = ctk.CTkComboBox(master=matriz_frame, values=list(COD),command=combobox_callback)
            self.codigo_sistema_1.place(x=270, y=70)
         

            labem_Ma_codigo_2 = ctk.CTkLabel(master=matriz_frame, text = 'Escolha o segundo codigo do sistema ', font=('Century Gothic bold',16), text_color='#fff').place(x=220,y=210)
            nome_sistema_2 = ctk.CTkComboBox(master=matriz_frame,values=['']).place(x=270, y=320)
            def combobox_callback(choice):
                if (choice):
                    #FILTRANDO OS PERFIS DE CADA SISTEMA
                    dataframeMatriz1 = pd.read_excel('.\sistemaEscola.xlsx', sheet_name='perfilSistema')
                    nomematriz1 = dataframeMatriz1.loc[dataframeMatriz1['CODIGO']==choice,'NOME']
                    self.nome_sistema_2 = ctk.CTkComboBox(master=matriz_frame,values=list(nomematriz1))
                    self.nome_sistema_2.place(x=270, y=320)
            label_sistema_2 = ctk.CTkLabel(master=matriz_frame, text = 'Escolha o perfil do sistema 2 ', font=('Century Gothic bold',16), text_color='#fff').place(x=252,y=285)
            self.codigo_sistema_2 = ctk.CTkComboBox(master=matriz_frame, values=list(COD),command=combobox_callback)
            self.codigo_sistema_2.place(x=270, y=245)
            
            
            def back():
                #removendo frame
                matriz_frame.pack_forget()

                #devolvendo frame da tela inicial
                inicial_frame.pack(side=RIGHT)
            voltar = ctk.CTkButton(master=matriz_frame, text='VOLTAR', font=('Century Gothic bold',16), text_color='#fff',command= back ).place(x=20, y=350)
            salvar = ctk.CTkButton(master=matriz_frame, text='SALVAR', font=('Century Gothic bold',16), text_color='#fff', fg_color='green',hover_color="#014B05", command= self.salvaMatriz).place(x=545, y=350)
            

        def tela_perfil_user():
            #remover tela inicial
            inicial_frame.pack_forget()

            #criando tela de cadastro de sistema
            perfilU_frame = ctk.CTkFrame(master=self, width=700, height= 400)
            perfilU_frame.pack(side=RIGHT)

            titulo = ctk.CTkLabel(master=perfilU_frame, text = 'Cadastre os perfis dos usuarios', font=('Century Gothic bold',16),  text_color='gray').place(x=20,y=10)
            
            #recuperando os valores do banco de dados
            dataframe = pd.read_excel('.\sistemaEscola.xlsx', sheet_name='Sistema')
            COD = dataframe.loc[:,'CODIGO']
            
            label = ctk.CTkLabel(master=perfilU_frame, text= 'Digite o CPF', font=('Century Gothic bold',16), text_color='#fff').place(x=290,y=35)
            self.cpf=ctk.CTkEntry(master=perfilU_frame, placeholder_text= 'xxx.xxx.xxx-xx', width=200)
            self.cpf.place(x=240,y=70)
            labem_Ma_codigo_1 = ctk.CTkLabel(master=perfilU_frame, text = 'Escolha o codigo do sistema ', font=('Century Gothic bold',16), text_color='#fff').place(x=252,y=110)
            self.nome_sistema = ctk.CTkComboBox(master=perfilU_frame,values=[''])
            self.nome_sistema.place(x=270, y=220)
            def combobox_callback(choice):
                #FILTRANDO OS PERFIS DE CADA SISTEMA
                dataframeMatriz1 = pd.read_excel('.\sistemaEscola.xlsx', sheet_name='perfilSistema')
                nomematriz1 = dataframeMatriz1.loc[dataframeMatriz1['CODIGO']==choice,'NOME']

                self.nome_sistema = ctk.CTkComboBox(master=perfilU_frame,values=list(nomematriz1))
                self.nome_sistema.place(x=270, y=220)

            label_sistema_1 = ctk.CTkLabel(master=perfilU_frame, text = 'Escolha o perfil do sistema', font=('Century Gothic bold',16), text_color='#fff').place(x=252,y=185)            
            self.codigo_sistema = ctk.CTkComboBox(master=perfilU_frame, values=list(COD),command=combobox_callback)
            self.codigo_sistema.place(x=270, y=145)
            
            def back():
                #removendo frame
                perfilU_frame.pack_forget()

                #devolvendo frame da tela inicial
                inicial_frame.pack(side=RIGHT)
            
            voltar = ctk.CTkButton(master=perfilU_frame, text='VOLTAR', font=('Century Gothic bold',16), text_color='#fff',command= back ).place(x=20, y=350)
            
            salvar = ctk.CTkButton(master=perfilU_frame, text='SALVAR', font=('Century Gothic bold',16), text_color='#fff', fg_color='green',hover_color="#014B05", command= self.salvarUser ).place(x=545, y=350)           
            
        #BOTÃOS DA TELA INICIAL 
        cadastroSistema = ctk.CTkButton(master=inicial_frame,text='Cadastros dos Sistemas',font=('Century Gothic bold',16), text_color='#fff', width=290, command=tela_sistemas).place(x=10,y=70)
        cadastroPerfis = ctk.CTkButton(master=inicial_frame,text='Cadastros dos perfis do Sistemas',font=('Century Gothic bold',16), text_color='#fff', width=290, command=tela_perfil).place(x=10,y=150)
        cadastroSMatriz = ctk.CTkButton(master=inicial_frame,text='Cadastros da matriz SOD',font=('Century Gothic bold',16), text_color='#fff', width=290, command=tela_matriz).place(x=400,y=70)
        cadastroPerfiluser = ctk.CTkButton(master=inicial_frame,text='Cadastros dos Perfils de usuarios',font=('Century Gothic bold',16), text_color='#fff', width=290, command=tela_perfil_user).place(x=400,y=150)

if __name__=="__main__":
    app = App()
    app.mainloop()